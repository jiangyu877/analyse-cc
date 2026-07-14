import csv
import math
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


class TabularDataError(ValueError):
    pass


@dataclass(frozen=True)
class TabularDataset:
    columns: tuple[str, ...]
    rows: tuple[dict[str, str], ...]


def normalize_header(value):
    value = str(value or "").strip().lower()
    value = re.sub(r"[\s\-/]+", "_", value)
    return value.strip("_")


def _cell_text(value):
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _decode_csv(data):
    if b"\x00" in data or data.startswith((b"MZ", b"\x7fELF")):
        raise TabularDataError("CSV 文件内容无效")
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise TabularDataError("CSV 文件必须使用 UTF-8 或 GB18030 编码")


def read_tabular(filename, data, max_rows=10000):
    filename = (filename or "").strip()
    if not filename or Path(filename).name != filename:
        raise TabularDataError("文件名不安全")
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        raise TabularDataError("仅支持 CSV 和 XLSX 文件")
    if not data:
        raise TabularDataError("不能上传空文件")

    workbook = None
    if extension == ".csv":
        reader = csv.DictReader(StringIO(_decode_csv(data)))
        if not reader.fieldnames:
            raise TabularDataError("数据集缺少表头")
        raw_columns = list(reader.fieldnames)
        raw_rows = reader
    else:
        if not data.startswith(b"PK"):
            raise TabularDataError("文件内容不是有效 XLSX")
        try:
            with zipfile.ZipFile(BytesIO(data)) as archive:
                members = archive.infolist()
                total_size = sum(item.file_size for item in members)
                if len(members) > 256 or total_size > 32 * 1024 * 1024:
                    raise TabularDataError("XLSX 解压内容过大")
                for item in members:
                    if item.filename == "xl/sharedStrings.xml" and item.file_size > 8 * 1024 * 1024:
                        raise TabularDataError("XLSX 共享文本内容过大")
                    if item.filename == "xl/styles.xml" and item.file_size > 2 * 1024 * 1024:
                        raise TabularDataError("XLSX 样式内容过大")
                    if (
                        item.filename.lower().endswith(".xml")
                        and "/worksheets/" not in item.filename.lower()
                        and item.file_size > 8 * 1024 * 1024
                    ):
                        raise TabularDataError("XLSX 元数据内容过大")
                if any(".." in Path(item.filename).parts for item in members):
                    raise TabularDataError("XLSX 文件结构不安全")
                member_map = {item.filename.lstrip("/"): item for item in members}
                manifest_info = member_map.get("[Content_Types].xml")
                if not manifest_info or manifest_info.file_size > 1024 * 1024:
                    raise TabularDataError("XLSX 内容类型清单无效")
                manifest = ElementTree.fromstring(archive.read(manifest_info))
                limits = {
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml": 8 * 1024 * 1024,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml": 2 * 1024 * 1024,
                }
                for override in manifest.findall("{http://schemas.openxmlformats.org/package/2006/content-types}Override"):
                    limit = limits.get(override.attrib.get("ContentType"))
                    part = member_map.get(override.attrib.get("PartName", "").lstrip("/"))
                    if limit and part and part.file_size > limit:
                        raise TabularDataError("XLSX 共享文本或样式内容过大")
            workbook = load_workbook(
                BytesIO(data), read_only=True, data_only=True, keep_links=False
            )
            worksheet = workbook.active
            values = worksheet.iter_rows(values_only=True)
            raw_columns = list(next(values))
            raw_rows = (
                {raw_columns[index]: row[index] if index < len(row) else None
                 for index in range(len(raw_columns))}
                for row in values
            )
        except TabularDataError:
            if workbook is not None:
                workbook.close()
            raise
        except StopIteration as exc:
            if workbook is not None:
                workbook.close()
            raise TabularDataError("数据集缺少表头") from exc
        except Exception as exc:
            if workbook is not None:
                workbook.close()
            raise TabularDataError("XLSX 文件无法读取，请检查文件是否损坏") from exc

    try:
        columns = tuple(normalize_header(column) for column in raw_columns)
        if any(not column for column in columns):
            raise TabularDataError("表头不能为空")
        if len(columns) != len(set(columns)):
            raise TabularDataError("表头存在重复列")
        if len(columns) > 100:
            raise TabularDataError("数据集最多支持 100 列")

        rows = []
        for scanned, raw_row in enumerate(raw_rows, start=1):
            row = {
                columns[index]: _cell_text(raw_row.get(raw_columns[index]))
                for index in range(len(columns))
            }
            if any(row.values()):
                rows.append(row)
            if len(rows) > max_rows or scanned > max_rows * 2:
                raise TabularDataError(f"单次最多导入 {max_rows} 行")
        if not rows:
            raise TabularDataError("数据集没有有效数据行")
        return TabularDataset(columns=columns, rows=tuple(rows))
    finally:
        if workbook is not None:
            workbook.close()


def remap_columns(dataset, aliases, required):
    lookup = {}
    for canonical, names in aliases.items():
        for name in (canonical, *names):
            lookup[normalize_header(name)] = canonical

    mapped_columns = {}
    for column in dataset.columns:
        canonical = lookup.get(column)
        if canonical:
            if canonical in mapped_columns.values():
                raise TabularDataError(f"字段 {canonical} 被重复映射")
            mapped_columns[column] = canonical
    missing = [field for field in required if field not in mapped_columns.values()]
    if missing:
        raise TabularDataError("缺少必填列：" + "、".join(missing))

    return tuple({mapped_columns[key]: value for key, value in row.items() if key in mapped_columns}
                 for row in dataset.rows)
